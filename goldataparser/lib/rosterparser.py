# # # # #
# Project: TurningParser
# File: 
# Created by: Nathan Healea
# Description:
# 
# # # # #

import os
import sys
from openpyxl import load_workbook
import string

class RosterParser:
    punctuation = '''!()-[]{};:'"\,<>./?@#$%^&*_~ '''
    def __init__(self):
        self.C = []  # Consent
        self.G = []  # Grades
        self.file = ''
        self.wb = None
        self.ws = {}
        self.match = []
        self.filename = None
        self.filepath = None
        self.key = ['firstname', 'lastname', 'sid', 'email', 'consent', 'grade','score']



    # #
    # Opens a give workbook and index worksheets
    # Parameters
    #   file: path to file to be open
    # Output
    #  True: File was open and sheets parsed
    #  False: File not opened or sheets not parsed
    def open(self, file):
        # Defining process flag
        status = True

        # Saving file name
        head, tail = os.path.split(file)
        self.filename = os.path.splitext(tail)[0]
        self.filepath = head

        # Setting File string and load workbook
        self.file = file
        self.wb = load_workbook(self.file)

        if self.wb == None:  # Check if workbook was opened
            status = False
            print ("ERROR: wb was not open")
        else:
            # Gets names of sheets from workbook
            wsNames = self.wb.get_sheet_names()
            count = 0

            # Adds sheets names and sheet to dictionary
            for sheet in self.wb.worksheets:
                self.ws[wsNames[count]] = sheet
                count += 1

            if len(self.ws) <= 0:  # Check if sheets where indexed
                status = False
                print ("ERROR: Workbooks sheet not index ")

        return status


    # #
    # Srubs punctation form strings
    # text: a string that contains punctuation
    def scrub(self, text):
        newText = ""
        for char in text:
            if char not in self.punctuation:
                newText = newText + char
        newText = newText.lower()
        return newText

    # #
    # Parses out sheets of a file
    # Parameters
    #   sheet: Name of the sheet to be parse
    # Output
    #   none
    # End Condition
    #    contains array of students from the passed sheet
    def parse(self):

        # Activating consentor worksheet
        classlist = []

        for row in self.ws['M'].iter_rows():
                count = 0
                student = {}
                for cell in row:
                    if(self.key[count] == 'firstname' or self.key[count] == 'lastname') :
                        if(cell.value is not None):
                            student[self.key[count]] = self.scrub(str(cell.value))
                        else:
                            student[self.key[count]] = cell.value
                    else:
                        student[self.key[count]] = cell.value
                    count += 1

                classlist.append(student)
            
        classlist = self.consentReplace(classlist)
        return classlist
    
    # #
    # Replaces bad consent data with zero
    def consentReplace(self, classlist):
        for student in classlist:
            if 'consent' in student:
                if student['consent'] != 1:
                    student['consent'] = 0
            else:
                student['consent'] = 0
        return classlist

    # #
    # Closes current workbook and removes all data from object
    #
    def close(self):
        self.C = []  # Consent
        self.G = []  # Grades
        self.file = ''
        self.wb = None
        self.ws = {}
        self.match = []
        self.filename = None
        self.filepath = None
