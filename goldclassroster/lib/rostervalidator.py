import os
import sys
from termcolor import colored
from openpyxl import load_workbook


class RosterValidator:
    def __init__(self):
        self.C = []  # Consent
        self.G = []  # Grades
        self.file = ''
        self.wb = None
        self.ws = {}
        self.match = []
        self.filename = None
        self.filepath = None
        self.key = ['First', 'Last', 'SID', 'Email', 'Grade', 'Consent', 'Score']

    # #
    # Opens a give workbook and index worksheets
    # Parameters
    #   file: path to file to be open
    # Output
    #  True: File was open and sheets parsed
    #  False: File not opened or sheets not parsed
    def open(self, file):
        # Defining process flag
        status = True

        # Saving file name
        head, tail = os.path.split(file)
        self.filename = os.path.splitext(tail)[0]
        self.filepath = head

        # Setting File string and load workbook
        self.file = file
        self.wb = load_workbook(self.file)

        if self.wb == None:  # Check if workbook was opened
            status = False
            print ("ERROR: wb was not open")
        else:
            # Gets names of sheets from workbook
            wsNames = self.wb.get_sheet_names()
            count = 0

            # Adds sheets names and sheet to dictionary
            for sheet in self.wb.worksheets:
                self.ws[wsNames[count]] = sheet
                count += 1

            if len(self.ws) <= 0:  # Check if sheets where indexed
                status = False
                print ("ERROR: Workbooks sheet not index ")

        return status

    # #
    # Parses out sheets of a file
    # Parameters
    #   sheet: Name of the sheet to be parse
    # Output
    #   none
    # End Condition
    #    contains array of students from the passed sheet
    def parse(self, sheet):

        # Defining process status
        status = True

        # Activating consentor worksheet

        firstRow = False
        index = []

        for row in self.ws[sheet].iter_rows():
            if not firstRow:
                for cell in row:
                    index.append(cell.value)

                firstRow = True
            else:
                count = 0
                student = {}
                for cell in row:
                    student[index[count]] = cell.value
                    count += 1
                student['Match'] = False
                getattr(self, sheet).append(student)

    # #
    # Compares  Grades (G) list to Consent (C) List
    def compair(self):
        for gStudent in self.G:
            # print "\tgStudent " + str(gStudent['First']) + ' ' + str(gStudent['Last'])
            match = False
            for cStudent in self.C:
                # print "\t\tcStudent " + str(cStudent['First'])  + ' ' + str(cStudent['Last'])


                if gStudent.has_key('SID') and cStudent.has_key('SID'):
                    if gStudent['SID'] == cStudent['SID'] and (gStudent['SID'] is not None) and (cStudent['SID'] is not None):
                        # print colored('\t\t\t\t SID: ' + str(
                        #     gStudent['SID']) + ' -- ' + str(cStudent['SID']),
                        #               'green')
                        match = True
                        break

                if gStudent['First'] == cStudent['First'] and gStudent[
                    'Last'] == cStudent['Last'] and match == False:
                    # print colored(
                    #     '\t\t' + str(gStudent['First']) + ' ' + str(gStudent[
                    #         'Last']) + ' -- ' + str(cStudent['First']) + ' ' + str(cStudent[
                    #         'Last']), 'green')
                    match = True
                    break
            if match:
                gStudent['Match'] = True
                cStudent['Match'] = True
                self.match.append(self.combine([gStudent, cStudent]))

        print "      |-- Num Matched: " + str(len(self.match))

        self.log(self.G, "Grade-Mismatch")
        self.log(self.C, "Consent-Mistmatch")

    def combine(self, student_list):

        c = {}
        for student in student_list:

            for key in student:
                if not c.has_key(key):
                    c[key] = student[key]

        for key in self.key:
            if not c.has_key(key):
                c[key] = " "

        return c

    # #
    # Logs students from list that did not have a match
    def log(self, list, name):

        if os.path.exists(
                '{0}/{1}-{2}.log'.format(self.filepath, self.filename, name)):
            os.remove(
                '{0}/{1}-{2}.log'.format(self.filepath, self.filename, name))

        fo = open('{0}/{1}-{2}.log'.format(self.filepath, self.filename, name),
                  "w")

        # Log list
        for student in list:
            if student['Match'] == False:
                for key in student:
                    fo.write("{0} ".format(student[key]))
                fo.write("\n")

        fo.close()

    # #
    # Records matched students into work books.
    def record(self):

        if "M" not in self.wb.get_sheet_names():
            self.wb.create_sheet(title="M")

        row = 1
        col = 1
        for student in self.match:
            self.wb["M"].cell(row=row, column=1).value = student['First']
            self.wb["M"].cell(row=row, column=2).value = student['Last']
            self.wb["M"].cell(row=row, column=3).value = student['SID']
            self.wb["M"].cell(row=row, column=4).value = student['Email']
            self.wb["M"].cell(row=row, column=5).value = student['Consent']
            self.wb["M"].cell(row=row, column=6).value = student['Grade']
            self.wb["M"].cell(row=row, column=7).value = student['Score']
            row += 1

        # Saves
        self.wb.save(filename=self.file)

    def close(self):
        self.C = []  # Consent
        self.G = []  # Grades
        self.file = ''
        self.wb = None
        self.ws = {}
        self.match = []
        self.filename = None
        self.filepath = None
